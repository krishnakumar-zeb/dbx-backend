"""XLSX Service – mask PII in Excel workbooks cell-by-cell."""
from fastapi import UploadFile
from typing import Dict
import openpyxl
import io
import re

from services.BaseService import BaseService
from utility.exceptions import FileValidationException, DocumentProcessingException


class XLSXService(BaseService):

    def _validate(self, document: UploadFile) -> None:
        fn = (document.filename or "").lower()
        if not fn.endswith(".xlsx"):
            raise FileValidationException("File must be an .xlsx file")

    def _extract_text(self, raw: bytes, filename: str) -> str:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
            parts = []
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    parts.extend(
                        str(cell.value) for cell in row if cell.value is not None
                    )
            return " ".join(parts)
        except Exception as e:
            raise DocumentProcessingException(f"XLSX parse failed: {e}")

    def _build_masked_output(self, raw, mapping, anonymized_text, out_path):
        """Replace PII in each cell of the workbook."""
        try:
            wb = openpyxl.load_workbook(io.BytesIO(raw))
            # Build original->tag from diff approach
            orig_text = self._extract_text(raw, "")
            replacements = self._build_replacement_pairs(orig_text, anonymized_text)

            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            val = str(cell.value)
                            masked = self._apply_replacements(val, replacements)
                            if masked != val:
                                cell.value = masked
            wb.save(out_path)
        except Exception as e:
            raise DocumentProcessingException(f"XLSX masking failed: {e}")

    @staticmethod
    def _build_replacement_pairs(original: str, anonymized: str) -> Dict[str, str]:
        """
        Compare original and anonymized text to extract {original_value: tag} pairs.
        Uses a simple diff approach: find tags in anonymized text and map them
        back to the corresponding spans in the original.
        """
        pairs: Dict[str, str] = {}
        tag_pattern = re.compile(r"<[A-Z_]+_\d+>")
        tags = list(tag_pattern.finditer(anonymized))
        if not tags:
            return pairs

        # Walk through both texts to align
        anon_pos = 0
        orig_pos = 0
        for match in tags:
            tag_start = match.start()
            tag_end = match.end()
            tag = match.group()

            # Characters before this tag should be the same in both
            prefix_len = tag_start - anon_pos
            orig_pos += prefix_len
            anon_pos = tag_end

            # Find the original value: it's the text in original at orig_pos
            # that was replaced by the tag. We need to figure out its length.
            # Look ahead to the next common text after the tag.
            if anon_pos < len(anonymized):
                # Find next non-tag text
                next_tag = tag_pattern.search(anonymized, anon_pos)
                if next_tag:
                    next_common = anonymized[anon_pos:next_tag.start()]
                else:
                    next_common = anonymized[anon_pos:]

                if next_common and next_common in original[orig_pos:]:
                    end_idx = original.index(next_common, orig_pos)
                    orig_value = original[orig_pos:end_idx]
                    if orig_value.strip():
                        pairs[orig_value] = tag
                    orig_pos = end_idx
                else:
                    # Fallback: skip ahead
                    orig_pos += len(tag)
            else:
                # Tag is at the end
                orig_value = original[orig_pos:]
                if orig_value.strip():
                    pairs[orig_value] = tag
                orig_pos = len(original)

        return pairs

    @staticmethod
    def _apply_replacements(text: str, replacements: Dict[str, str]) -> str:
        """Apply all replacement pairs to text, longest match first."""
        for orig in sorted(replacements, key=len, reverse=True):
            text = text.replace(orig, replacements[orig])
        return text

